from __future__ import annotations

from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DETAIL_HEADERS = [
    "Title",
    "Internal Path",
    "Target",
    "Impact area",
    "Suggestion",
    "Manual Review Required",
    "Reasoning",
]

AZURE_HEADERS = [
    "ID",
    "State",
    "Tags",
    "Work Item Type",
    "Title",
    "Internal Path",
    "Target",
    "Impact area",
    "Suggestion",
    "Incidents Repeats",
    "Review type",
    "Developer",
    "Iteration Path",
]


RULE_TITLE_MAP = {
    "Hardcode": "Hard Code",
    "Parametrizable": "Parametrizable",
    "Manejo de errores (RunAfter)": "Manejo de errores",
    "Retrasos": "Retrasos",
    "Retrasos (Delay y Wait)": "Retrasos",
    "Nomenclatura de actividades": "Nombre de Actividades",
    "Nomenclatura de variables": "Nombre de Variables",
    "Prefijos variables / parámetros": "Nombre de Argumentos",
    "Nomenclatura de flujos": "Nombre Módulos",
    "Condición IF": "Condición IF",
    "Descripciones de acciones": "Descripciones",
    "Comentarios descriptivos": "Descripciones",
    "Scopes vacíos o con una sola acción": "Scope y Region",
    "Switch caso predeterminado vacío": "Caso predeterminado",
    "Límite de anidación en condición": "Condición IF",
    "Id de requerimiento": "Id de requerimiento",
    "Descripciones de flujos": "Descripciones",
}


def _write_sheet(ws, headers, rows, col_widths=None):
    thin = Side(style="thin", color="D9DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    if col_widths:
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"


def _azure_title(finding: dict[str, Any]) -> str:
    rule_name = str(finding.get("rule_name", "") or "").strip()
    return RULE_TITLE_MAP.get(rule_name, rule_name)


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _flow_name(finding: dict[str, Any]) -> str:
    return _clean_text(finding.get("flow_name"))


def _short_target(finding: dict[str, Any]) -> str:
    return _clean_text(finding.get("target_pretty") or finding.get("target"))


def _flow_activity_path(finding: dict[str, Any]) -> str:
    flow = _flow_name(finding)
    short_target = _short_target(finding)

    if flow and short_target:
        return f"{flow} / {short_target}"
    if flow:
        return flow
    return short_target


def _detail_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    rows = []

    for finding in findings:
        rows.append(
            [
                _azure_title(finding),
                _flow_activity_path(finding),
                _short_target(finding),
                _clean_text(finding.get("impact_area")),
                _clean_text(finding.get("suggestion")),
                _clean_text(finding.get("manual_review_required")),
                _clean_text(finding.get("reasoning")),
            ]
        )

    return rows


def _azure_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "internal_paths": [],
            "targets": [],
            "count": 0,
        }
    )

    for finding in findings:
        title = _azure_title(finding)
        impact_area = _clean_text(finding.get("impact_area"))
        suggestion = _clean_text(finding.get("suggestion"))

        key = (title, impact_area, suggestion)
        item = grouped[key]

        short_target = _short_target(finding)
        flow_activity = _flow_activity_path(finding)

        if short_target and short_target not in item["internal_paths"]:
            item["internal_paths"].append(short_target)

        if flow_activity and flow_activity not in item["targets"]:
            item["targets"].append(flow_activity)

        item["count"] += int(finding.get("repeat_count", 1) or 1)

    rows = []

    for (title, impact_area, suggestion), item in grouped.items():
        rows.append(
            [
                "",
                "Active",
                "",
                "Code Review",
                title,
                "\n".join(item["internal_paths"]),
                "\n".join(item["targets"]),
                impact_area,
                suggestion,
                item["count"],
                "",
                "",
                "",
            ]
        )

    rows.sort(key=lambda row: (row[4], row[7], row[5]))
    return rows


def export_review_to_xlsx(
    out_path,
    *,
    findings: list[dict[str, Any]],
    project_id: str,
    usage_data: dict[str, Any] | None = None,
) -> None:
    wb = Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    ws_details = wb.create_sheet("details")
    ws_azure = wb.create_sheet("azure_ready")

    _write_sheet(
        ws_details,
        DETAIL_HEADERS,
        _detail_rows(findings),
        col_widths={
            1: 28,
            2: 72,
            3: 48,
            4: 24,
            5: 95,
            6: 22,
            7: 70,
        },
    )

    _write_sheet(
        ws_azure,
        AZURE_HEADERS,
        _azure_rows(findings),
        col_widths={
            1: 10,
            2: 14,
            3: 18,
            4: 18,
            5: 28,
            6: 55,
            7: 72,
            8: 24,
            9: 95,
            10: 18,
            11: 16,
            12: 18,
            13: 20,
        },
    )

    wb.save(out_path)